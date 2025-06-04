from flask import (
    Flask,
    request,
    jsonify,
    send_file,
    render_template,
    send_from_directory,
)
from flask_cors import CORS
from dotenv import load_dotenv
import os
import openpyxl
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Initialize Flask app with correct static folder configuration
app = Flask(
    __name__,
    static_folder=".",  # Set static folder to root directory
    static_url_path="",  # Remove /static prefix from URLs
    template_folder="templates",
)

# CORS configuration
CORS(
    app,
    resources={
        r"/*": {
            "origins": [
                "https://www.tringtriadmarketing.com",
                "http://localhost:3000",
                "https://v6j.90b.mytemp.website",
            ],
            "methods": ["GET", "POST", "OPTIONS"],
            "allow_headers": ["Content-Type", "Authorization"],
            "supports_credentials": True,
        }
    },
)


def handle_excel_file(submission):
    excel_file_path = os.path.join(
        os.path.dirname(__file__), "contact_submissions.xlsx"
    )

    try:
        # Check if file exists and is valid
        if os.path.exists(excel_file_path):
            try:
                workbook = openpyxl.load_workbook(excel_file_path)
                worksheet = workbook["Submissions"]
            except Exception as e:
                logger.error(f"Error loading existing Excel file: {str(e)}")
                # Create new workbook if file is corrupted
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Submissions"
                worksheet.append(["Timestamp", "Name", "Email", "Company", "Message"])
        else:
            # Create new workbook if file doesn't exist
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Submissions"
            worksheet.append(["Timestamp", "Name", "Email", "Company", "Message"])

        # Add new submission
        worksheet.append(
            [
                datetime.now().isoformat(),
                submission["name"],
                submission["email"],
                submission.get("company", "N/A"),
                submission.get("message", "N/A"),
            ]
        )

        # Save workbook
        workbook.save(excel_file_path)
        logger.info("Successfully saved submission to Excel file")
    except Exception as e:
        logger.error(f"Error saving to Excel: {str(e)}")
        raise


@app.route("/test", methods=["GET"])
def test():
    return jsonify({"status": "ok", "message": "Flask app is running"})


@app.route("/send-email", methods=["POST", "OPTIONS"])
def send_email():
    logger.debug(f"Received request: {request.method}")
    logger.debug(f"Request headers: {request.headers}")
    logger.debug(f"Request data: {request.get_data()}")

    if request.method == "OPTIONS":
        return "", 200

    try:
        data = request.get_json()
        logger.debug(f"Parsed JSON data: {data}")

        if not data or not data.get("name") or not data.get("email"):
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Missing required fields (name, email)",
                    }
                ),
                400,
            )

        # Save to Excel file
        try:
            handle_excel_file(data)
        except Exception as e:
            logger.error(f"Error saving to Excel: {str(e)}")
            # Continue with email sending even if Excel save fails

        # Email configuration
        msg = MIMEMultipart("alternative")
        msg["From"] = os.getenv("SMTP_FROM_EMAIL", "admin@tringtriadmarketing.com")
        msg["To"] = os.getenv("RECIPIENT_EMAIL", "admin@tringtriadmarketing.com")
        msg["Subject"] = f'New Contact Form Submission from {data["name"]}'

        # Email body - plain text version
        text = f"""Name: {data['name']}
Email: {data['email']}
Company: {data.get('company', 'N/A')}
Message: {data.get('message', 'No message provided')}"""

        # Email body - HTML version
        html = f"""<html>
<body>
<p><strong>Name:</strong> {data['name']}</p>
<p><strong>Email:</strong> {data['email']}</p>
<p><strong>Company:</strong> {data.get('company', 'N/A')}</p>
<p><strong>Message:</strong> {data.get('message', 'No message provided')}</p>
</body>
</html>"""

        # Attach both versions
        msg.attach(MIMEText(text, "plain"))
        msg.attach(MIMEText(html, "html"))

        try:
            # Get SMTP configuration with defaults
            smtp_host = os.getenv(
                "SMTP_HOST", "mail.tringtriadmarketing.com"
            )  # Update with your hosting provider's SMTP server
            smtp_port = int(
                os.getenv("SMTP_PORT", "587")
            )  # Default to 587 if not specified
            smtp_user = os.getenv("SMTP_USER", "admin@tringtriadmarketing.com")
            smtp_password = os.getenv(
                "SMTP_PASSWORD", ""
            )  # You'll need to set this in your .env file
            smtp_secure = os.getenv("SMTP_SECURE", "true").lower() == "true"

            # Validate required SMTP settings
            if not all([smtp_host, smtp_user, smtp_password]):
                logger.error("Missing required SMTP configuration")
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "Email server configuration is incomplete",
                        }
                    ),
                    500,
                )

            # Send email
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.set_debuglevel(1)  # Enable debug output
                try:
                    # Initial EHLO
                    server.ehlo()

                    # Always use STARTTLS for Gmail
                    if smtp_secure:
                        if not server.has_extn("STARTTLS"):
                            raise smtplib.SMTPException(
                                "STARTTLS extension not supported by server"
                            )
                        server.starttls()
                        server.ehlo()  # Re-identify ourselves over TLS connection

                    # Verify AUTH support after STARTTLS
                    if not server.has_extn("AUTH"):
                        raise smtplib.SMTPException(
                            "AUTH extension not supported by server"
                        )

                    # Login with credentials
                    server.login(smtp_user, smtp_password)

                    # Send the message
                    server.send_message(msg)
                    logger.info("Email sent successfully")

                except smtplib.SMTPAuthenticationError as e:
                    logger.error(f"SMTP Authentication failed: {str(e)}")
                    return (
                        jsonify(
                            {
                                "success": False,
                                "message": "Email authentication failed. Please check your credentials.",
                            }
                        ),
                        500,
                    )
                except smtplib.SMTPException as e:
                    logger.error(f"SMTP error: {str(e)}")
                    return (
                        jsonify(
                            {
                                "success": False,
                                "message": f"Failed to send email: {str(e)}",
                            }
                        ),
                        500,
                    )

            return jsonify({"success": True, "message": "Email sent successfully!"})
        except Exception as e:
            logger.error(f"Error sending email: {str(e)}")
            return (
                jsonify(
                    {
                        "success": False,
                        "message": "Failed to send email. Please try again.",
                    }
                ),
                500,
            )
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return (
            jsonify({"success": False, "message": "An unexpected error occurred"}),
            500,
        )


@app.route("/download-submissions", methods=["GET"])
def download_submissions():
    excel_file_path = os.path.join(
        os.path.dirname(__file__), "contact_submissions.xlsx"
    )

    if not os.path.exists(excel_file_path):
        return jsonify({"success": False, "message": "No submissions found"}), 404

    return send_file(
        excel_file_path, as_attachment=True, download_name="contact_submissions.xlsx"
    )


@app.route("/download/submission")
def submission_download():
    return render_template("submission_download.html")


# Serve static files
@app.route("/<path:path>")
def serve_static(path):
    return send_from_directory(".", path)


# Root route
@app.route("/")
def index():
    return send_from_directory(".", "index.html")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 3000)))
